''' Пространственное распределение давления по зоне выстрела
из статьи  DOI:10.3390/ma14081878
'''
from math import *
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook

#******** различные функции Гаусса
def gausPressGleb(r, s):
    return 1 / (s * (2 * pi) ** 0.5) * exp(- (r ** 2 / (2 * s ** 2)))

def gausPressArtical(r, R):    # https://doi.org/10.3390/ma14081878
    return exp(- (r ** 2 / (2 * R ** 2)))

def gausPressNormal(r):
    return 1 / (2 * pi) ** 0.5 * exp(- (r ** 2 / (2 * 1 ** 2)))

def nogaus(r, a, b):
    return exp (-2*(r / a) ** b)
#**************************

# Исходные данные
diametr = 0.14   # диаметр пятна, cm
Power = 4   # Энергия, Джоули
tay = 60   # длительность импульса, ns

gamma = 5/3     # адиабатическая что-то там
alpha = 0.25    # подгоночный коэффициент, связанная с долей ионизации

# табличные данные, для титана Ti6Al4V брал от сюда doi: 10.1063/1.1303508
p_w, p_tit = 1, 4.42  #density of water and titan, gram/cm^3 
D_w, D_tit = 1.483, 4.987 # 1.465, 4.987 - or speed of sound (https://www.matweb.com/search/datasheet_print.aspx?matguid=b350a789eda946c6b86a3e4d3c577b39)| 5.2 - speed of shockwave in water and in titan, km/s

# shock impedance
Z1 = p_w * D_w * 1e5
Z2 = p_tit * D_tit * 1e5
Z = 2 * Z2 * Z1 / (Z1 + Z2)

R = 0.7   #spot radius in mm
Intensity0 = Power / (tay * np.pi * (2*R/10) ** 2 / 4) # GW/cm^2
P_peak = 0.01 * ((alpha * Z * Intensity0) / (2 * alpha + 3)) ** 0.5  # pressure in GPa по Фабро


s = 2*R / 2 / (2 * log(2)) ** 0.5 # в данном случае 2R это полуширина в Функции Гаусса


# Точность
Xincrement = 0.001

P = []
r = R * (2 * log(10000)) ** 0.5  #радиус при котором P = 0.01*P_tay сотая часть от максимума
r_argument = np.arange(-r, r, Xincrement)

# Площади прямоугольника и Гаусса
S1 = 2*R * P_peak
S2 = 0

def grafPress(func, s):
    global S2
    S2_part = 0
    for i in r_argument:
        pressure = func(i, s) #* P_peak
        P.append(pressure)
        S2 = Xincrement * pressure + S2
        if s >= i >= -s:
            S2_part += Xincrement * pressure
    
    return S2_part / S2 * 100 # при нормальном распределении около 68,2%


print(grafPress(gausPressGleb, s))


P = [i * S1 / S2 for i in P] # приведение к одной площади

# print(P_peak, 'v', (max(P)), 'NewP / Pmax =', max(P)/P_peak ) #(P[int(len(P)/2)-1]))
print(S1, S2, S1 / S2)

# оператор повышения контрастности нечеткого множества
def operINT(PressureMassive):
    operP = []
    PressureMassive = [i / max(PressureMassive) for i in PressureMassive]
    for i in range(len(PressureMassive)):
        if PressureMassive[i] < 0.5:
            operP.append(2*PressureMassive[i]**2)
        else:
            operP.append(1 - 2*(1-PressureMassive[i])**2)

    return operP

# сколько раз модифицировать исходного Гаусса
n = 0
Pmodif = P
for i in range(n):
    Pmodif = operINT(Pmodif)

if n != 0:
    Pmodif = [Pmodif[i] * P_peak for i in range(len(P))]

print(max(Pmodif))

#*********** график **********
x_name = 'Расстояние от точки до центра пятна, мм'
y_name = 'Давление, ГПа'

plt.plot(r_argument, Pmodif, color = 'brown')
# plt.plot(r_argument, P, color = 'green')

r_argument2 = [-R, -R, R, R]
P_average = [0, P_peak, P_peak, 0]
plt.plot(r_argument2, P_average)

plt.plot([-2, 2], [max(P) / 2 for i in range(2)], color = 'black')

plt.grid()

plt.xlabel(x_name)
plt.ylabel(y_name)

plt.show()

''''
#********* создание таблицы ***********
wb = Workbook()
sheet = wb.active

def printColumn(rw, cl, massive):
    for i in range(len(massive)):
        c = sheet.cell(row = rw+i, column = cl)
        c.value = massive[i]

c_name = sheet['E1']
c_name.value = 'Исходные данные:'
c_name = sheet['F2']
c_name.value = 'P_peak, ГПа'
c_name = sheet['F3']
c_name.value = 'R, мм'
c_name = sheet['G2']
c_name.value = P_peak
c_name = sheet['G3']
c_name.value = R

c_name = sheet['A1']
c_name.value = y_name
c_name = sheet['B1']
c_name.value = x_name

printColumn(2, 1, P)
printColumn(2, 2, r)

# n = P   #для пербора
# for j in range(2):
#     for i in range(len(r)):
#         c = sheet.cell(row = i+2, column = j+1)
#         c.value = n[i]

#     n = r

wb.save('data20.xlsx')    #Creat or just update file in your working directory
'''